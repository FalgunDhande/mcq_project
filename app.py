
from flask import Flask, render_template, request, redirect, url_for, flash, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, login_user, login_required, logout_user, current_user, UserMixin
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta
from sqlalchemy import inspect, text
import os, csv, io, random, json
from openpyxl import load_workbook
from reportlab.pdfgen import canvas

app = Flask(__name__, instance_relative_config=True)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'devsecret')
os.makedirs(app.instance_path, exist_ok=True)

# --------- DB config ---------
db_url = os.environ.get('DATABASE_URL')
if db_url:
    if db_url.startswith("postgres://"):
        db_url = db_url.replace("postgres://", "postgresql+psycopg://", 1)
    elif db_url.startswith("postgresql://") and "+psycopg" not in db_url:
        db_url = db_url.replace("postgresql://", "postgresql+psycopg://", 1)
else:
    db_url = 'sqlite:///' + os.path.join(app.instance_path, 'quiz.db')

app.config['SQLALCHEMY_DATABASE_URI'] = db_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

app.config['UPLOAD_FOLDER'] = os.path.join(os.path.dirname(__file__), 'uploads')
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

db = SQLAlchemy(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'

# ---------------- Models ----------------
class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(200), nullable=False)
    role = db.Column(db.String(20), default='user')
    coins = db.Column(db.Integer, default=0)
    badges = db.Column(db.String(200), default='')
    streak = db.Column(db.Integer, default=0)
    def set_password(self, pw): self.password_hash = generate_password_hash(pw)
    def check_password(self, pw): return check_password_hash(self.password_hash, pw)

class Subject(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), unique=True, nullable=False)

class Chapter(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    subject_id = db.Column(db.Integer, db.ForeignKey('subject.id'), nullable=False)
    name = db.Column(db.String(120), nullable=False)
    subject = db.relationship('Subject', backref='chapters')

class Quiz(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    duration_minutes = db.Column(db.Integer, default=10)
    negative_marking = db.Column(db.Float, default=0.0)
    marks_per_question = db.Column(db.Float, default=1.0)
    start_time = db.Column(db.DateTime, nullable=True)
    end_time = db.Column(db.DateTime, nullable=True)
    questions = db.relationship('Question', backref='quiz', cascade="all, delete-orphan")

class Question(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    quiz_id = db.Column(db.Integer, db.ForeignKey('quiz.id'), nullable=False)
    subject_id = db.Column(db.Integer, db.ForeignKey('subject.id'), nullable=True)
    chapter_id = db.Column(db.Integer, db.ForeignKey('chapter.id'), nullable=True)
    text = db.Column(db.Text, nullable=False)
    option_a = db.Column(db.Text, nullable=False)
    option_b = db.Column(db.Text, nullable=False)
    option_c = db.Column(db.Text, nullable=False)
    option_d = db.Column(db.Text, nullable=False)
    correct_option = db.Column(db.String(1), nullable=False)
    difficulty = db.Column(db.String(10), default='Medium')
    qtype = db.Column(db.String(20), default='Concept')
    subject = db.relationship('Subject')
    chapter = db.relationship('Chapter')

class Assignment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    quiz_id = db.Column(db.Integer, db.ForeignKey('quiz.id'), nullable=False)
    attempts_limit = db.Column(db.Integer, default=1)
    cooldown_days = db.Column(db.Integer, default=0)
    user = db.relationship('User', backref='assignments')
    quiz = db.relationship('Quiz', backref='assignments')

class Attempt(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    quiz_id = db.Column(db.Integer, db.ForeignKey('quiz.id'), nullable=False)
    score = db.Column(db.Float, default=0.0)
    started_at = db.Column(db.DateTime, default=datetime.utcnow)
    submitted_at = db.Column(db.DateTime, nullable=True)
    active = db.Column(db.Boolean, default=True)
    user = db.relationship('User', backref='attempts')
    quiz = db.relationship('Quiz', backref='attempts')

class AttemptAnswer(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    attempt_id = db.Column(db.Integer, db.ForeignKey('attempt.id'), nullable=False)
    question_id = db.Column(db.Integer, db.ForeignKey('question.id'), nullable=False)
    selected_option = db.Column(db.String(1), nullable=True)
    is_correct = db.Column(db.Boolean, default=False)
    marks_earned = db.Column(db.Float, default=0.0)
    flagged = db.Column(db.Boolean, default=False)
    note = db.Column(db.Text, default='')
    attempt = db.relationship('Attempt', backref='answers')
    question = db.relationship('Question')

class Message(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    to_user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    from_user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    body = db.Column(db.Text, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class BlueprintSpec(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False)
    spec_json = db.Column(db.Text, nullable=False)

@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))

def auto_migrate_user_table():
    insp = inspect(db.engine)
    cols = {c['name'] for c in insp.get_columns('user')}
    dialect = db.engine.dialect.name
    tbl = '"user"' if dialect == 'postgresql' else 'user'
    ine = 'IF NOT EXISTS ' if dialect == 'postgresql' else ''
    stmts = []
    if 'coins' not in cols:   stmts.append(f'ALTER TABLE {tbl} ADD COLUMN {ine} coins INTEGER DEFAULT 0')
    if 'badges' not in cols:  stmts.append(f"ALTER TABLE {tbl} ADD COLUMN {ine} badges VARCHAR(200) DEFAULT ''")
    if 'streak' not in cols:  stmts.append(f'ALTER TABLE {tbl} ADD COLUMN {ine} streak INTEGER DEFAULT 0')
    for s in stmts:
        try:
            db.session.execute(text(s)); db.session.commit()
        except Exception:
            db.session.rollback()

with app.app_context():
    db.create_all()
    auto_migrate_user_table()
    if not User.query.filter_by(username='admin').first():
        admin = User(username='admin', role='admin'); admin.set_password('admin123')
        db.session.add(admin); db.session.commit()

# ---------- helpers ----------
def get_or_create_subject(name):
    if not name: return None
    s = Subject.query.filter(db.func.lower(Subject.name)==name.lower()).first()
    if not s: s = Subject(name=name); db.session.add(s); db.session.commit()
    return s

def get_or_create_chapter(subject, name):
    if not name or not subject: return None
    c = Chapter.query.filter(db.func.lower(Chapter.name)==name.lower(), Chapter.subject_id==subject.id).first()
    if not c: c = Chapter(name=name, subject_id=subject.id); db.session.add(c); db.session.commit()
    return c

# -------------- auth --------------
@app.route('/login', methods=['GET','POST'])
def login():
    if request.method == 'POST':
        u = User.query.filter_by(username=request.form.get('username','').strip()).first()
        if u and u.check_password(request.form.get('password','')):
            login_user(u); return redirect(url_for('index'))
        flash('Invalid credentials','error')
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user(); return redirect(url_for('login'))

# -------------- dashboards --------------
@app.route('/')
@login_required
def index():
    if current_user.role in ('admin','teacher','moderator'):
        stats = dict(
            users=User.query.count(),
            quizzes=Quiz.query.count(),
            questions=Question.query.count(),
            attempts=Attempt.query.count()
        )
        top_users = db.session.query(User.username, db.func.sum(Attempt.score).label('total')).join(Attempt, Attempt.user_id==User.id).group_by(User.id).order_by(db.desc('total')).limit(10).all()
        active = Attempt.query.filter_by(active=True, submitted_at=None).order_by(Attempt.started_at.desc()).all()
        blueprints = BlueprintSpec.query.order_by(BlueprintSpec.id.desc()).all()
        return render_template('admin_dashboard.html', stats=stats, quizzes=Quiz.query.all(), users=User.query.all(), blueprints=blueprints, leaderboard=top_users, active=active)
    assignments = Assignment.query.filter_by(user_id=current_user.id).all()
    attempts = Attempt.query.filter_by(user_id=current_user.id).order_by(Attempt.started_at.desc()).all()
    msgs = Message.query.filter_by(to_user_id=current_user.id).order_by(Message.created_at.desc()).limit(20).all()
    return render_template('user_dashboard.html', assignments=assignments, attempts=attempts, msgs=msgs)

# -------------- Add User --------------
@app.route('/admin/add_user', methods=['POST'])
@login_required
def admin_add_user():
    if current_user.role not in ('admin','teacher'):
        return redirect(url_for('index'))
    username = request.form.get('username','').strip()
    password = request.form.get('password','').strip()
    role = request.form.get('role','user')
    if not username or not password:
        flash('Username & password required', 'error'); return redirect(url_for('index'))
    if User.query.filter_by(username=username).first():
        flash('User already exists', 'error'); return redirect(url_for('index'))
    u = User(username=username, role=role); u.set_password(password); db.session.add(u); db.session.commit()
    flash('User created', 'success'); return redirect(url_for('index'))

# -------------- Create Quiz --------------
@app.route('/admin/create_quiz', methods=['POST'])
@login_required
def admin_create_quiz():
    if current_user.role not in ('admin','teacher'): return redirect(url_for('index'))
    title = request.form.get('title','').strip() or 'Quiz'
    start = request.form.get('start_time','').strip()
    end = request.form.get('end_time','').strip()
    qz = Quiz(title=title, duration_minutes=int(request.form.get('duration',10)),
              negative_marking=float(request.form.get('negative',0)),
              marks_per_question=float(request.form.get('marks',1)),
              start_time=(datetime.fromisoformat(start) if start else None),
              end_time=(datetime.fromisoformat(end) if end else None))
    db.session.add(qz); db.session.commit()
    flash('Quiz created','success'); return redirect(url_for('index'))

# -------------- Delete Quiz --------------
@app.route('/admin/quiz/delete/<int:quiz_id>', methods=['POST'])
@login_required
def admin_delete_quiz(quiz_id):
    if current_user.role not in ('admin','teacher'): return redirect(url_for('index'))
    quiz = db.session.get(Quiz, quiz_id)
    if not quiz: flash('Quiz not found','error'); return redirect(url_for('index'))
    db.session.query(AttemptAnswer).filter(AttemptAnswer.attempt_id.in_(db.session.query(Attempt.id).filter_by(quiz_id=quiz_id))).delete(synchronize_session=False)
    db.session.query(Attempt).filter_by(quiz_id=quiz_id).delete(synchronize_session=False)
    db.session.query(Assignment).filter_by(quiz_id=quiz_id).delete(synchronize_session=False)
    db.session.delete(quiz); db.session.commit()
    flash('Quiz deleted','success'); return redirect(url_for('index'))

# -------------- Upload Questions --------------
def add_question_row(quiz, row):
    subj = get_or_create_subject((row.get('subject') or '').strip()) if row.get('subject') else None
    chap = get_or_create_chapter(subj, (row.get('chapter') or '').strip()) if (subj and row.get('chapter')) else None
    q = Question(quiz_id=quiz.id,
                 subject_id=(subj.id if subj else None),
                 chapter_id=(chap.id if chap else None),
                 text=str(row.get('question') or row.get('Question')),
                 option_a=str(row.get('option_a') or row.get('A')),
                 option_b=str(row.get('option_b') or row.get('B')),
                 option_c=str(row.get('option_c') or row.get('C')),
                 option_d=str(row.get('option_d') or row.get('D')),
                 correct_option=str(row.get('correct_option') or row.get('Answer') or 'A')[:1].upper(),
                 difficulty=str(row.get('difficulty') or 'Medium').title(),
                 qtype=str(row.get('qtype') or 'Concept').title())
    db.session.add(q)

@app.route('/admin/upload_questions', methods=['POST'])
@login_required
def admin_upload_questions():
    if current_user.role not in ('admin','teacher'): return redirect(url_for('index'))
    quiz_id = request.form.get('quiz_id', type=int)
    if not quiz_id: flash('Select quiz','error'); return redirect(url_for('index'))
    quiz = db.session.get(Quiz, quiz_id)
    f = request.files.get('file')
    if not f or f.filename=='': flash('Choose a file','error'); return redirect(url_for('index'))
    ext = f.filename.lower().rsplit('.',1)[-1]
    path = os.path.join(app.config['UPLOAD_FOLDER'], f.filename); f.save(path)
    added=0
    try:
        if ext=='csv':
            with open(path, newline='', encoding='utf-8') as cf:
                reader = csv.DictReader(cf)
                for row in reader:
                    add_question_row(quiz, row); added+=1
        elif ext=='xlsx':
            wb = load_workbook(path); sh = wb.active
            headers=[c.value.strip().lower() if isinstance(c.value,str) else '' for c in next(sh.iter_rows(min_row=1, max_row=1))]
            idx={h:i for i,h in enumerate(headers)}
            def gv(r,k): 
                i=idx.get(k); 
                return (r[i].value if i is not None else None)
            for r in sh.iter_rows(min_row=2):
                row = {k:gv(r,k) for k in headers}
                add_question_row(quiz, row); added+=1
        else:
            flash('Use CSV/XLSX','error'); return redirect(url_for('index'))
        db.session.commit(); flash(f'Added {added} questions','success')
    except Exception as e:
        db.session.rollback(); flash(f'Upload error: {e}','error')
    return redirect(url_for('index'))

# -------------- Blueprints --------------
@app.route('/admin/blueprint/new', methods=['POST'])
@login_required
def create_blueprint():
    if current_user.role not in ('admin','teacher'): return redirect(url_for('index'))
    name = request.form.get('name','Blueprint')
    rules = request.form.get('rules','{"rules":[]}')
    bp = BlueprintSpec(name=name, spec_json=rules); db.session.add(bp); db.session.commit()
    flash('Blueprint saved','success'); return redirect(url_for('index'))

@app.route('/admin/blueprint/generate/<int:bp_id>', methods=['POST'])
@login_required
def generate_from_blueprint(bp_id):
    if current_user.role not in ('admin','teacher'): return redirect(url_for('index'))
    bp = db.session.get(BlueprintSpec, bp_id)
    if not bp: flash('Blueprint not found','error'); return redirect(url_for('index'))
    spec = json.loads(bp.spec_json)
    quiz = Quiz(title=f"Auto: {bp.name}", duration_minutes=int(request.form.get('duration',10)))
    db.session.add(quiz); db.session.commit()
    total=0
    for rule in spec.get('rules', []):
        q = Question.query
        if rule.get('subject'):
            subj = Subject.query.filter(db.func.lower(Subject.name)==rule['subject'].lower()).first()
            if subj: q = q.filter_by(subject_id=subj.id)
        if rule.get('chapter'):
            chap = Chapter.query.filter(db.func.lower(Chapter.name)==rule['chapter'].lower()).first()
            if chap: q = q.filter_by(chapter_id=chap.id)
        if rule.get('difficulty'):
            q = q.filter_by(difficulty=rule['difficulty'].title())
        pool = q.all(); random.shuffle(pool)
        for item in pool[:int(rule.get('count',0))]:
            clone = Question(quiz_id=quiz.id, text=item.text, option_a=item.option_a, option_b=item.option_b, option_c=item.option_c, option_d=item.option_d,
                             correct_option=item.correct_option, difficulty=item.difficulty, qtype=item.qtype, subject_id=item.subject_id, chapter_id=item.chapter_id)
            db.session.add(clone); total += 1
    db.session.commit(); flash(f'Generated quiz with {total} questions','success'); return redirect(url_for('index'))

# -------------- Assign Quiz --------------
@app.route('/admin/assign', methods=['POST'])
@login_required
def assign_quiz():
    if current_user.role not in ('admin','teacher','moderator'): return redirect(url_for('index'))
    quiz_id = request.form.get('quiz_id', type=int)
    limit = request.form.get('limit', type=int) or 1
    cooldown = request.form.get('cooldown', type=int) or 0
    ids = [int(i) for i in request.form.getlist('user_ids') if i.isdigit()]
    for uid in ids:
        ex = Assignment.query.filter_by(user_id=uid, quiz_id=quiz_id).first()
        if ex:
            ex.attempts_limit = limit; ex.cooldown_days = cooldown
        else:
            db.session.add(Assignment(user_id=uid, quiz_id=quiz_id, attempts_limit=limit, cooldown_days=cooldown))
    db.session.commit(); flash(f'Assigned to {len(ids)} users','success'); return redirect(url_for('index'))

# -------------- Start / Autosave / Submit --------------
@app.route('/quiz/<int:quiz_id>/start')
@login_required
def start_quiz(quiz_id):
    quiz = db.session.get(Quiz, quiz_id)
    if not quiz: flash('Quiz not found','error'); return redirect(url_for('index'))
    now = datetime.utcnow()
    if quiz.start_time and now < quiz.start_time: flash('Quiz not started yet','error'); return redirect(url_for('index'))
    if quiz.end_time and now > quiz.end_time: flash('Quiz time window ended','error'); return redirect(url_for('index'))
    ass = Assignment.query.filter_by(user_id=current_user.id, quiz_id=quiz_id).first()
    if current_user.role not in ('admin','teacher'):
        if not ass: flash('Quiz not assigned','error'); return redirect(url_for('index'))
        used = Attempt.query.filter_by(user_id=current_user.id, quiz_id=quiz_id).count()
        if used >= ass.attempts_limit: flash('Attempt limit reached','error'); return redirect(url_for('index'))
        last = Attempt.query.filter_by(user_id=current_user.id, quiz_id=quiz_id).order_by(Attempt.started_at.desc()).first()
        if ass.cooldown_days and last and (now - last.started_at).days < ass.cooldown_days:
            flash('Cooldown active','error'); return redirect(url_for('index'))
    at = Attempt(user_id=current_user.id, quiz_id=quiz_id, started_at=now, active=True)
    db.session.add(at); db.session.commit()
    questions = Question.query.filter_by(quiz_id=quiz.id).all(); random.shuffle(questions)
    deadline = now + timedelta(minutes=quiz.duration_minutes)
    return render_template('take_quiz.html', quiz=quiz, questions=questions, deadline_iso=deadline.isoformat()+'Z', attempt_id=at.id)

@app.route('/attempt/<int:attempt_id>/autosave', methods=['POST'])
@login_required
def autosave(attempt_id):
    at = db.session.get(Attempt, attempt_id)
    if not at or at.user_id != current_user.id: return ('', 403)
    qid = request.form.get('question_id', type=int)
    sel = request.form.get('selected','')
    flag = request.form.get('flag','') == '1'
    note = request.form.get('note','')
    aa = AttemptAnswer.query.filter_by(attempt_id=at.id, question_id=qid).first()
    if not aa:
        aa = AttemptAnswer(attempt_id=at.id, question_id=qid); db.session.add(aa)
    if sel: aa.selected_option = sel[:1].upper()
    aa.flagged = flag
    if note: aa.note = note[:300]
    db.session.commit(); return ('', 204)

@app.route('/quiz/<int:quiz_id>/submit', methods=['POST'])
@login_required
def submit_quiz(quiz_id):
    quiz = db.session.get(Quiz, quiz_id)
    attempt_id = request.form.get('attempt_id', type=int)
    at = db.session.get(Attempt, attempt_id)
    if not at: return redirect(url_for('index'))
    questions = Question.query.filter_by(quiz_id=quiz.id).all()
    score=0.0
    for q in questions:
        a = AttemptAnswer.query.filter_by(attempt_id=at.id, question_id=q.id).first()
        sel = a.selected_option if a and a.selected_option else request.form.get(f'q_{q.id}', None)
        is_correct = (sel is not None and sel[:1].upper()==q.correct_option)
        marks = quiz.marks_per_question if is_correct else (-quiz.negative_marking if sel is not None else 0.0)
        score += marks
        if not a:
            a = AttemptAnswer(attempt_id=at.id, question_id=q.id); db.session.add(a)
        a.selected_option = (sel[:1].upper() if sel else None); a.is_correct=is_correct; a.marks_earned=marks
    at.score=score; at.submitted_at=datetime.utcnow(); at.active=False
    current_user.coins = (current_user.coins or 0) + max(0, int(score))
    if score >= 30:
        badges = set(filter(None, (current_user.badges or '').split(','))); badges.add('Topper'); current_user.badges=','.join(sorted(badges))
    db.session.commit()
    return redirect(url_for('review_attempt', attempt_id=at.id))

# -------------- Review / PDF / Export --------------
@app.route('/attempt/<int:attempt_id>/review')
@login_required
def review_attempt(attempt_id):
    at = db.session.get(Attempt, attempt_id)
    if not at: flash('Attempt not found','error'); return redirect(url_for('index'))
    if current_user.role not in ('admin','teacher') and at.user_id != current_user.id:
        flash('Unauthorized','error'); return redirect(url_for('index'))
    answers = AttemptAnswer.query.filter_by(attempt_id=at.id).all()
    per_subject, per_chapter = {}, {}
    items=[]
    for a in answers:
        q = db.session.get(Question, a.question_id)
        subj = q.subject.name if q.subject else 'General'
        chap = q.chapter.name if q.chapter else '-'
        ps = per_subject.setdefault(subj, {"total":0,"correct":0,"wrong":0,"marks":0.0})
        pc = per_chapter.setdefault((subj,chap), {"total":0,"correct":0,"wrong":0,"marks":0.0})
        ps["total"]+=1; pc["total"]+=1
        if a.is_correct: ps["correct"]+=1; pc["correct"]+=1
        else:
            if a.selected_option is not None: ps["wrong"]+=1; pc["wrong"]+=1
        ps["marks"] += a.marks_earned; pc["marks"] += a.marks_earned
        items.append({"q":q,"a":a,"subject":subj,"chapter":chap})
    return render_template('review.html', attempt=at, items=items, per_subject=per_subject, per_chapter=per_chapter)

@app.route('/attempt/<int:attempt_id>/report.pdf')
@login_required
def report_pdf(attempt_id):
    at = db.session.get(Attempt, attempt_id)
    if not at: return ('Not found',404)
    buf = io.BytesIO()
    c = canvas.Canvas(buf)
    c.setFont("Helvetica-Bold", 14); c.drawString(50, 800, f"Result Report - {at.quiz.title}")
    c.setFont("Helvetica", 11); c.drawString(50, 780, f"User: {at.user.username}"); c.drawString(50, 765, f"Score: {at.score}")
    y = 740
    answers = AttemptAnswer.query.filter_by(attempt_id=at.id).all()
    for a in answers[:40]:
        q = db.session.get(Question, a.question_id)
        line = f"Q{q.id}: Ans={a.selected_option or '-'} | Correct={q.correct_option} | Marks={a.marks_earned}"
        c.drawString(50, y, line[:100]); y -= 14
        if y < 60: c.showPage(); y = 800
    c.showPage(); c.save(); buf.seek(0)
    return send_file(buf, mimetype='application/pdf', as_attachment=True, download_name='report.pdf')

@app.route('/admin/export/questions.csv')
@login_required
def export_questions():
    if current_user.role not in ('admin','teacher'): return redirect(url_for('index'))
    buf = io.StringIO(); w = csv.writer(buf)
    w.writerow(['quiz','subject','chapter','difficulty','qtype','question','A','B','C','D','correct'])
    for q in Question.query.all():
        w.writerow([q.quiz.title, q.subject.name if q.subject else '', q.chapter.name if q.chapter else '', q.difficulty, q.qtype, q.text, q.option_a, q.option_b, q.option_c, q.option_d, q.correct_option])
    buf.seek(0); return send_file(io.BytesIO(buf.getvalue().encode('utf-8')), mimetype='text/csv', as_attachment=True, download_name='questions.csv')

@app.route('/admin/export/attempts.csv')
@login_required
def export_attempts():
    if current_user.role not in ('admin','teacher'): return redirect(url_for('index'))
    buf = io.StringIO(); w = csv.writer(buf)
    w.writerow(['user','quiz','score','started','submitted'])
    for a in Attempt.query.all():
        w.writerow([a.user.username, a.quiz.title, a.score, a.started_at, a.submitted_at])
    buf.seek(0); return send_file(io.BytesIO(buf.getvalue().encode('utf-8')), mimetype='text/csv', as_attachment=True, download_name='attempts.csv')

if __name__ == '__main__':
    app.run(debug=True)
